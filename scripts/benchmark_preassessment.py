"""Parse CDP preassessment workbooks into calibration benchmark outputs.

The script reads client preassessment XLSX files from local paths and writes
summaries under outputs/. It intentionally does not copy source workbooks or
embed evidence text, so confidential client files stay outside the repository.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


LEVELS = ("D", "A", "M", "L")
QUESTION_RE = re.compile(r"^\d+(?:\.\d+){1,3}[a-z]?$", re.IGNORECASE)


@dataclass
class BenchmarkRow:
    company: str
    workbook: str
    sheet: str
    row_number: int
    module: str
    category: str
    question: str
    sector_flag: str
    d_flag: str
    a_flag: str
    m_flag: str
    l_flag: str
    d_score: float
    d_max: float
    a_score: float
    a_max: float
    m_score: float
    m_max: float
    l_score: float
    l_max: float
    total_gap: float
    largest_gap_level: str
    dataset_d_max: float | None = None
    dataset_a_max: float | None = None
    dataset_m_max: float | None = None
    dataset_l_max: float | None = None
    denominator_match: str = "not_checked"


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        cleaned = re.sub(r"[^0-9.\-]", "", text(value))
        try:
            return float(cleaned) if cleaned else 0.0
        except ValueError:
            return 0.0


def safe_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", value)[:31]
    return cleaned or "Sheet"


def company_from_workbook(path: Path, first_sheet_title: str = "") -> str:
    name = path.stem
    for token in ("_2025", " 2025", "-2025", "_CDP", " CDP", "-CDP"):
        if token in name:
            return name.split(token)[0].strip("_ -")
    return first_sheet_title or name


def find_header(ws: Any) -> dict[str, int] | None:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=False), start=1):
        labels = [text(cell.value) for cell in row]
        joined = " ".join(labels)
        if "Question Number" not in joined or "D(점수)" not in joined:
            continue
        indexes = {label: idx + 1 for idx, label in enumerate(labels) if label}
        required = ["Question Number", "D(점수)", "D(배점)", "A(점수)", "A(배점)", "M(점수)", "M(배점)", "L(점수)", "L(배점)"]
        if all(item in indexes for item in required):
            return {
                "header_row": row_number,
                "module": indexes.get("Module", 0),
                "category": indexes.get("Scoring Category", 0),
                "question": indexes["Question Number"],
                "sector": indexes.get("섹터문항 여부", 0),
                "d_flag": indexes.get("D", 0),
                "a_flag": indexes.get("A", 0),
                "m_flag": indexes.get("M", 0),
                "l_flag": indexes.get("L", 0),
                "d_score": indexes["D(점수)"],
                "d_max": indexes["D(배점)"],
                "a_score": indexes["A(점수)"],
                "a_max": indexes["A(배점)"],
                "m_score": indexes["M(점수)"],
                "m_max": indexes["M(배점)"],
                "l_score": indexes["L(점수)"],
                "l_max": indexes["L(배점)"],
            }
    return None


def candidate_sheets(path: Path) -> list[tuple[Any, dict[str, int]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    candidates: list[tuple[Any, dict[str, int]]] = []
    for ws in wb.worksheets:
        header = find_header(ws)
        if header:
            candidates.append((ws, header))
    return candidates


def parse_workbook(path: Path) -> tuple[list[BenchmarkRow], dict[str, Any]]:
    candidates = candidate_sheets(path)
    if not candidates:
        return [], {"workbook": path.name, "status": "no_header"}

    best_ws = None
    best_header = None
    best_count = -1
    for ws, header in candidates:
        count = 0
        for row_index in range(header["header_row"] + 1, ws.max_row + 1):
            qn = text(ws.cell(row_index, header["question"]).value)
            if QUESTION_RE.match(qn):
                count += 1
        if count > best_count:
            best_ws, best_header, best_count = ws, header, count

    assert best_ws is not None and best_header is not None
    company = company_from_workbook(path, text(best_ws.cell(1, 3).value))
    rows: list[BenchmarkRow] = []
    for row_index in range(best_header["header_row"] + 1, best_ws.max_row + 1):
        qn = text(best_ws.cell(row_index, best_header["question"]).value)
        if not QUESTION_RE.match(qn):
            continue
        values = {
            "d_score": to_number(best_ws.cell(row_index, best_header["d_score"]).value),
            "d_max": to_number(best_ws.cell(row_index, best_header["d_max"]).value),
            "a_score": to_number(best_ws.cell(row_index, best_header["a_score"]).value),
            "a_max": to_number(best_ws.cell(row_index, best_header["a_max"]).value),
            "m_score": to_number(best_ws.cell(row_index, best_header["m_score"]).value),
            "m_max": to_number(best_ws.cell(row_index, best_header["m_max"]).value),
            "l_score": to_number(best_ws.cell(row_index, best_header["l_score"]).value),
            "l_max": to_number(best_ws.cell(row_index, best_header["l_max"]).value),
        }
        gaps = {
            "D": max(values["d_max"] - values["d_score"], 0),
            "A": max(values["a_max"] - values["a_score"], 0),
            "M": max(values["m_max"] - values["m_score"], 0),
            "L": max(values["l_max"] - values["l_score"], 0),
        }
        largest_level = max(gaps, key=gaps.get)
        rows.append(
            BenchmarkRow(
                company=company,
                workbook=path.name,
                sheet=best_ws.title,
                row_number=row_index,
                module=text(best_ws.cell(row_index, best_header["module"]).value) if best_header["module"] else "",
                category=text(best_ws.cell(row_index, best_header["category"]).value) if best_header["category"] else "",
                question=qn,
                sector_flag=text(best_ws.cell(row_index, best_header["sector"]).value) if best_header["sector"] else "",
                d_flag=text(best_ws.cell(row_index, best_header["d_flag"]).value) if best_header["d_flag"] else "",
                a_flag=text(best_ws.cell(row_index, best_header["a_flag"]).value) if best_header["a_flag"] else "",
                m_flag=text(best_ws.cell(row_index, best_header["m_flag"]).value) if best_header["m_flag"] else "",
                l_flag=text(best_ws.cell(row_index, best_header["l_flag"]).value) if best_header["l_flag"] else "",
                total_gap=sum(gaps.values()),
                largest_gap_level=largest_level if gaps[largest_level] else "",
                **values,
            )
        )
    return rows, {"workbook": path.name, "status": "ok", "sheet": best_ws.title, "question_rows": len(rows)}


def load_dataset_denominators(path: Path) -> dict[str, dict[str, float]]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    mapping: dict[str, dict[str, float]] = {}
    for row in data.get("qualitativeRows", []):
        qn = text(row.get("questionNumber") or row.get("question_number"))
        denoms = row.get("denominators") or {}
        if qn:
            mapping[qn] = {level: to_number(denoms.get(level)) for level in LEVELS}
    return mapping


def attach_dataset_comparison(rows: list[BenchmarkRow], denominators: dict[str, dict[str, float]]) -> None:
    for row in rows:
        dataset = denominators.get(row.question)
        if not dataset:
            row.denominator_match = "question_not_in_dataset"
            continue
        row.dataset_d_max = dataset.get("D")
        row.dataset_a_max = dataset.get("A")
        row.dataset_m_max = dataset.get("M")
        row.dataset_l_max = dataset.get("L")
        pairs = [
            (row.d_max, row.dataset_d_max),
            (row.a_max, row.dataset_a_max),
            (row.m_max, row.dataset_m_max),
            (row.l_max, row.dataset_l_max),
        ]
        row.denominator_match = "match" if all(abs((a or 0) - (b or 0)) < 0.0001 for a, b in pairs) else "mismatch"


def workbook_summary(rows: list[BenchmarkRow], parse_statuses: list[dict[str, Any]]) -> list[list[Any]]:
    output: list[list[Any]] = [["Company", "Workbook", "Sheet", "Question Rows", "Total Gap", "Denominator Match", "Mismatch", "Question Not In Dataset"]]
    by_workbook: dict[str, list[BenchmarkRow]] = {}
    for row in rows:
        by_workbook.setdefault(row.workbook, []).append(row)
    for status in parse_statuses:
        wb_rows = by_workbook.get(status["workbook"], [])
        output.append(
            [
                wb_rows[0].company if wb_rows else "",
                status["workbook"],
                status.get("sheet", ""),
                len(wb_rows),
                round(sum(row.total_gap for row in wb_rows), 4),
                sum(1 for row in wb_rows if row.denominator_match == "match"),
                sum(1 for row in wb_rows if row.denominator_match == "mismatch"),
                sum(1 for row in wb_rows if row.denominator_match == "question_not_in_dataset"),
            ]
        )
    return output


def row_values(row: BenchmarkRow) -> list[Any]:
    data = asdict(row)
    return [data[key] for key in data]


def write_xlsx(rows: list[BenchmarkRow], statuses: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    for values in workbook_summary(rows, statuses):
        summary.append(values)

    parsed = wb.create_sheet("Parsed Scores")
    headers = list(asdict(rows[0]).keys()) if rows else list(BenchmarkRow.__dataclass_fields__.keys())
    parsed.append(headers)
    for row in rows:
        parsed.append(row_values(row))

    gaps = wb.create_sheet("Deduction Priority")
    gaps.append(["Company", "Question", "Module", "Category", "Total Gap", "Largest Gap Level", "D", "A", "M", "L", "Workbook", "Sheet", "Row"])
    for row in sorted(rows, key=lambda item: item.total_gap, reverse=True)[:200]:
        gaps.append(
            [
                row.company,
                row.question,
                row.module,
                row.category,
                row.total_gap,
                row.largest_gap_level,
                f"{row.d_score}/{row.d_max}",
                f"{row.a_score}/{row.a_max}",
                f"{row.m_score}/{row.m_max}",
                f"{row.l_score}/{row.l_max}",
                row.workbook,
                row.sheet,
                row.row_number,
            ]
        )

    mismatches = wb.create_sheet("Denominator Check")
    mismatches.append(["Company", "Question", "Result", "Benchmark D/A/M/L", "Dataset D/A/M/L", "Workbook", "Sheet", "Row"])
    for row in rows:
        if row.denominator_match == "match":
            continue
        mismatches.append(
            [
                row.company,
                row.question,
                row.denominator_match,
                f"{row.d_max}/{row.a_max}/{row.m_max}/{row.l_max}",
                f"{row.dataset_d_max}/{row.dataset_a_max}/{row.dataset_m_max}/{row.dataset_l_max}",
                row.workbook,
                row.sheet,
                row.row_number,
            ]
        )

    fill = PatternFill("solid", fgColor="EAF2F8")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
        for column in ws.columns:
            max_len = max(len(text(cell.value)) for cell in column[:100])
            ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 10), 45)

    wb.save(output_path)


def write_json(rows: list[BenchmarkRow], statuses: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "parseStatuses": statuses,
        "summary": workbook_summary(rows, statuses)[1:],
        "rows": [asdict(row) for row in rows],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def read_path_file(path: Path) -> list[Path]:
    return [Path(line.strip()) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def main() -> None:
    parser = argparse.ArgumentParser(description="Build CDP preassessment benchmark outputs.")
    parser.add_argument("files", nargs="*", type=Path, help="Preassessment XLSX files")
    parser.add_argument("--path-file", type=Path, help="UTF-8 text file containing one XLSX path per line")
    parser.add_argument("--dataset", type=Path, default=Path("data/cdp_2026_full_dataset.json"))
    parser.add_argument("--output-xlsx", type=Path, default=Path("outputs/benchmark/preassessment_benchmark.xlsx"))
    parser.add_argument("--output-json", type=Path, default=Path("outputs/benchmark/preassessment_benchmark.json"))
    args = parser.parse_args()

    paths = list(args.files)
    if args.path_file:
        paths.extend(read_path_file(args.path_file))
    if not paths:
        raise SystemExit("No input files supplied.")

    rows: list[BenchmarkRow] = []
    statuses: list[dict[str, Any]] = []
    for path in paths:
        parsed, status = parse_workbook(path)
        rows.extend(parsed)
        statuses.append(status)

    attach_dataset_comparison(rows, load_dataset_denominators(args.dataset))
    write_xlsx(rows, statuses, args.output_xlsx)
    write_json(rows, statuses, args.output_json)
    print(json.dumps({"rows": len(rows), "xlsx": str(args.output_xlsx), "json": str(args.output_json), "statuses": statuses}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
